from __future__ import annotations

import copy
import io
import json
import re
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from .database import AuditEventRecord, OrderRecord


class DomainError(ValueError):
    pass


def load_document(order: OrderRecord) -> dict[str, Any]:
    document = json.loads(order.document_json)
    for typology in document.get("typologies", []):
        for item in typology.get("glasses", []):
            if "quantity_final" not in item:
                item["quantity_original"] = item.get("quantity", 1)
                item["quantity_final"] = item.get("quantity", 1)
                item["measure_original"] = item.get("measure", "")
                item["measure_final"] = item.get("measure", "")
                item["description_original"] = item.get("description", "")
                item["description_final"] = item.get("description", "")
            item.setdefault("material_type", "glass")
            item.setdefault("observations", "")
            item.setdefault("status", "detected")
            item.setdefault("origin", "winmaker")
            item.setdefault("excluded", False)
            item.setdefault("reviewed", True)
        for accessory in typology.get("accessories", []):
            accessory.setdefault("observations", "")
            if accessory.get("status") == "quantity_modified":
                accessory["status"] = "modified"
    return document


def save_document(order: OrderRecord, document: dict[str, Any]) -> None:
    order.document_json = json.dumps(document, ensure_ascii=False, separators=(",", ":"))
    order.updated_at = datetime.now(timezone.utc)


def public_order(order: OrderRecord, include_document: bool = True) -> dict[str, Any]:
    payload = {
        "id": order.id,
        "client_name": order.client_name,
        "source_filename": order.source_filename,
        "status": order.status,
        "created_at": order.created_at.isoformat(),
        "updated_at": order.updated_at.isoformat(),
    }
    if include_document:
        payload["document"] = load_document(order)
    return payload


def find_typology(document: dict[str, Any], typology_id: str) -> dict[str, Any]:
    for typology in document["typologies"]:
        if typology["id"] == typology_id:
            return typology
    raise DomainError("No se encontró la tipología.")


def find_accessory(typology: dict[str, Any], accessory_id: str) -> dict[str, Any]:
    for accessory in typology["accessories"]:
        if accessory["id"] == accessory_id:
            return accessory
    raise DomainError("No se encontró el accesorio.")


def find_glass(typology: dict[str, Any], glass_id: str) -> dict[str, Any]:
    for glass in typology["glasses"]:
        if glass["id"] == glass_id:
            return glass
    raise DomainError("No se encontró el vidrio o tela.")


def add_audit(
    session: Session,
    order: OrderRecord,
    event_type: str,
    *,
    typology: dict[str, Any] | None = None,
    entity_type: str | None = None,
    entity_id: str | None = None,
    field: str | None = None,
    original: Any = None,
    previous: Any = None,
    final: Any = None,
    payload: dict[str, Any] | None = None,
) -> None:
    event = AuditEventRecord(
        order_id=order.id,
        event_type=event_type,
        typology_id=typology["id"] if typology else None,
        page=typology["page"] if typology else None,
        entity_type=entity_type,
        entity_id=entity_id,
        field=field,
        original_value=None if original is None else json.dumps(original, ensure_ascii=False),
        previous_value=None if previous is None else json.dumps(previous, ensure_ascii=False),
        final_value=None if final is None else json.dumps(final, ensure_ascii=False),
        payload_json=json.dumps(payload or {}, ensure_ascii=False),
    )
    session.add(event)


def update_accessory_quantity(
    session: Session,
    order: OrderRecord,
    typology_id: str,
    accessory_id: str,
    quantity: int | float,
) -> dict[str, Any]:
    if Decimal(str(quantity)) < 0:
        raise DomainError("La cantidad no puede ser negativa.")
    document = load_document(order)
    typology = find_typology(document, typology_id)
    accessory = find_accessory(typology, accessory_id)
    previous = accessory["quantity_final"]
    accessory["quantity_final"] = quantity
    if accessory["origin"] == "manual":
        accessory["status"] = "manual"
    elif (
        Decimal(str(quantity)) == Decimal(str(accessory["quantity_original"]))
        and not accessory.get("observations")
    ):
        accessory["status"] = "detected"
    else:
        accessory["status"] = "modified"
    accessory["reviewed"] = True
    document["modifications"].append(
        {
            "id": str(uuid.uuid4()),
            "type": "quantity_modified",
            "entity_type": "accessory",
            "entity_id": accessory_id,
            "typology_id": typology_id,
            "page": typology["page"],
            "field": "quantity",
            "original_value": accessory["quantity_original"],
            "previous_value": previous,
            "final_value": quantity,
            "date": datetime.now(timezone.utc).isoformat(),
        }
    )
    add_audit(
        session,
        order,
        "quantity_modified",
        typology=typology,
        entity_type="accessory",
        entity_id=accessory_id,
        field="quantity",
        original=accessory["quantity_original"],
        previous=previous,
        final=quantity,
    )
    save_document(order, document)
    return accessory


def toggle_accessory_exclusion(
    session: Session,
    order: OrderRecord,
    typology_id: str,
    accessory_id: str,
    excluded: bool,
) -> dict[str, Any]:
    document = load_document(order)
    typology = find_typology(document, typology_id)
    accessory = find_accessory(typology, accessory_id)
    previous = accessory["excluded"]
    accessory["excluded"] = excluded
    accessory["status"] = "excluded" if excluded else (
        "manual" if accessory["origin"] == "manual" else
        "modified"
        if Decimal(str(accessory["quantity_final"])) != Decimal(str(accessory["quantity_original"]))
        or bool(accessory.get("observations"))
        else "detected"
    )
    accessory["reviewed"] = True
    add_audit(
        session,
        order,
        "accessory_excluded" if excluded else "accessory_restored",
        typology=typology,
        entity_type="accessory",
        entity_id=accessory_id,
        field="excluded",
        original=False,
        previous=previous,
        final=excluded,
    )
    save_document(order, document)
    return accessory


def add_manual_accessory(
    session: Session,
    order: OrderRecord,
    typology_id: str,
    code: str,
    quantity: int | float,
    detail: str,
    observations: str = "",
) -> dict[str, Any]:
    code = re.sub(r"\s+", "", code).upper()
    if not code or not detail.strip():
        raise DomainError("Código y detalle son obligatorios.")
    if Decimal(str(quantity)) < 0:
        raise DomainError("La cantidad no puede ser negativa.")
    document = load_document(order)
    typology = find_typology(document, typology_id)
    accessory = {
        "id": str(uuid.uuid4()),
        "code_original": code,
        "code": code,
        "quantity_original": None,
        "quantity_final": quantity,
        "detail_original": detail.strip(),
        "detail_final": detail.strip(),
        "status": "manual",
        "origin": "manual",
        "confidence": "high",
        "excluded": False,
        "reviewed": True,
        "warnings": [],
        "observations": observations.strip(),
        "source": {"page": typology["page"], "row_id": None, "original_text": None},
        "stock_ref": None,
        "location_ref": None,
    }
    typology["accessories"].append(accessory)
    add_audit(
        session,
        order,
        "accessory_added",
        typology=typology,
        entity_type="accessory",
        entity_id=accessory["id"],
        final=accessory,
    )
    save_document(order, document)
    return accessory


def update_accessory_observations(
    session: Session,
    order: OrderRecord,
    typology_id: str,
    accessory_id: str,
    observations: str,
) -> dict[str, Any]:
    document = load_document(order)
    typology = find_typology(document, typology_id)
    accessory = find_accessory(typology, accessory_id)
    previous = accessory.get("observations", "")
    accessory["observations"] = observations.strip()
    if accessory["origin"] == "manual":
        accessory["status"] = "manual"
    elif accessory["observations"]:
        accessory["status"] = "modified"
    elif Decimal(str(accessory["quantity_final"])) == Decimal(str(accessory["quantity_original"])):
        accessory["status"] = "detected"
    accessory["reviewed"] = True
    add_audit(
        session,
        order,
        "accessory_observations_modified",
        typology=typology,
        entity_type="accessory",
        entity_id=accessory_id,
        field="observations",
        previous=previous,
        final=accessory["observations"],
    )
    save_document(order, document)
    return accessory


def update_glass(
    session: Session,
    order: OrderRecord,
    typology_id: str,
    glass_id: str,
    *,
    quantity: int | float,
    measure: str,
    description: str,
    observations: str,
) -> dict[str, Any]:
    if Decimal(str(quantity)) < 0:
        raise DomainError("La cantidad no puede ser negativa.")
    measure = re.sub(r"\s*[xX×]\s*", " x ", measure.strip())
    match = re.fullmatch(r"(\d+)\s+x\s+(\d+)", measure)
    if not match:
        raise DomainError("La medida debe tener el formato ancho x alto, por ejemplo 430 x 682.")
    if not description.strip():
        raise DomainError("La descripción es obligatoria.")
    document = load_document(order)
    typology = find_typology(document, typology_id)
    glass = find_glass(typology, glass_id)
    previous = {
        "quantity": glass["quantity_final"],
        "measure": glass["measure_final"],
        "description": glass["description_final"],
        "observations": glass.get("observations", ""),
    }
    glass["quantity_final"] = quantity
    glass["measure_final"] = measure
    glass["width"] = int(match.group(1))
    glass["height"] = int(match.group(2))
    glass["description_final"] = description.strip()
    glass["observations"] = observations.strip()
    changed = (
        Decimal(str(quantity)) != Decimal(str(glass["quantity_original"]))
        or measure != glass["measure_original"]
        or glass["description_final"] != glass["description_original"]
        or bool(glass["observations"])
    )
    glass["status"] = "manual" if glass["origin"] == "manual" else ("modified" if changed else "detected")
    glass["reviewed"] = True
    final = {
        "quantity": glass["quantity_final"],
        "measure": glass["measure_final"],
        "description": glass["description_final"],
        "observations": glass["observations"],
    }
    document["modifications"].append(
        {
            "id": str(uuid.uuid4()),
            "type": "glass_modified",
            "entity_type": "glass",
            "entity_id": glass_id,
            "typology_id": typology_id,
            "page": typology["page"],
            "original_value": {
                "quantity": glass["quantity_original"],
                "measure": glass["measure_original"],
                "description": glass["description_original"],
            },
            "previous_value": previous,
            "final_value": final,
            "date": datetime.now(timezone.utc).isoformat(),
        }
    )
    add_audit(
        session,
        order,
        "glass_modified",
        typology=typology,
        entity_type="glass",
        entity_id=glass_id,
        original={
            "quantity": glass["quantity_original"],
            "measure": glass["measure_original"],
            "description": glass["description_original"],
        },
        previous=previous,
        final=final,
    )
    save_document(order, document)
    return glass


def toggle_glass_exclusion(
    session: Session,
    order: OrderRecord,
    typology_id: str,
    glass_id: str,
    excluded: bool,
) -> dict[str, Any]:
    document = load_document(order)
    typology = find_typology(document, typology_id)
    glass = find_glass(typology, glass_id)
    previous = glass["excluded"]
    glass["excluded"] = excluded
    changed = (
        Decimal(str(glass["quantity_final"])) != Decimal(str(glass["quantity_original"]))
        or glass["measure_final"] != glass["measure_original"]
        or glass["description_final"] != glass["description_original"]
        or bool(glass.get("observations"))
    )
    glass["status"] = "excluded" if excluded else (
        "manual" if glass["origin"] == "manual" else ("modified" if changed else "detected")
    )
    add_audit(
        session,
        order,
        "glass_excluded" if excluded else "glass_restored",
        typology=typology,
        entity_type="glass",
        entity_id=glass_id,
        field="excluded",
        original=False,
        previous=previous,
        final=excluded,
    )
    save_document(order, document)
    return glass


def add_manual_glass(
    session: Session,
    order: OrderRecord,
    typology_id: str,
    *,
    material_type: str,
    quantity: int | float,
    measure: str,
    description: str,
    observations: str,
) -> dict[str, Any]:
    if material_type not in ("glass", "mesh"):
        raise DomainError("El tipo de material no es válido.")
    if Decimal(str(quantity)) < 0:
        raise DomainError("La cantidad no puede ser negativa.")
    measure = re.sub(r"\s*[xX×]\s*", " x ", measure.strip())
    match = re.fullmatch(r"(\d+)\s+x\s+(\d+)", measure)
    if not match:
        raise DomainError("La medida debe tener el formato ancho x alto.")
    if not description.strip():
        raise DomainError("La descripción es obligatoria.")
    document = load_document(order)
    typology = find_typology(document, typology_id)
    glass = {
        "id": str(uuid.uuid4()),
        "source_row": None,
        "material_type": material_type,
        "quantity_original": None,
        "quantity_final": quantity,
        "measure_original": None,
        "measure_final": measure,
        "width": int(match.group(1)),
        "height": int(match.group(2)),
        "description_original": None,
        "description_final": description.strip(),
        "observations": observations.strip(),
        "status": "manual",
        "origin": "manual",
        "excluded": False,
        "reviewed": True,
        "source": {"page": typology["page"], "row_id": None, "original_text": None},
        "confidence": "high",
    }
    typology["glasses"].append(glass)
    add_audit(
        session,
        order,
        "glass_added",
        typology=typology,
        entity_type="glass",
        entity_id=glass["id"],
        final=glass,
    )
    save_document(order, document)
    return glass


def confirm_typology(session: Session, order: OrderRecord, typology_id: str) -> None:
    document = load_document(order)
    typology = find_typology(document, typology_id)
    unresolved = [
        accessory
        for accessory in typology["accessories"]
        if accessory["confidence"] == "low" and not accessory.get("reviewed")
    ]
    if unresolved:
        raise DomainError("Revise los accesorios marcados con baja confianza antes de confirmar.")
    typology["review"]["status"] = "confirmed"
    typology["review"]["confirmed_at"] = datetime.now(timezone.utc).isoformat()
    add_audit(session, order, "typology_confirmed", typology=typology)
    save_document(order, document)


def _natural_key(value: str) -> list[Any]:
    return [int(part) if part.isdigit() else part.casefold() for part in re.split(r"(\d+)", value)]


def glass_rows(order: OrderRecord) -> list[dict[str, Any]]:
    document = load_document(order)
    rows: list[dict[str, Any]] = []
    for typology in document["typologies"]:
        for glass in typology["glasses"]:
            if glass["excluded"]:
                continue
            rows.append(
                {
                    "Cliente": order.client_name,
                    "Tipología": typology["typology"]["value"],
                    "Detalle abertura": typology["detail"]["value"],
                    "Tipo": "Tela mosquitera" if glass["material_type"] == "mesh" else "Vidrio",
                    "Cantidad": glass["quantity_final"],
                    "Medida": glass["measure_final"],
                    "Descripción": glass["description_final"],
                    "Observaciones": glass.get("observations", ""),
                }
            )
    return rows


def consolidated_accessories(order: OrderRecord) -> list[dict[str, Any]]:
    document = load_document(order)
    grouped: dict[str, dict[str, Any]] = {}
    resolutions = document.get("conflict_resolutions", {})
    for typology in document["typologies"]:
        for accessory in typology["accessories"]:
            if accessory["excluded"]:
                continue
            code = accessory["code"]
            item = grouped.setdefault(
                code,
                {
                    "Código": code,
                    "Cantidad Total": Decimal("0"),
                    "descriptions": [],
                    "winmaker_original": Decimal("0"),
                    "winmaker_modified": Decimal("0"),
                    "manual": Decimal("0"),
                },
            )
            quantity = Decimal(str(accessory["quantity_final"]))
            item["Cantidad Total"] += quantity
            if accessory["detail_final"] and accessory["detail_final"] not in item["descriptions"]:
                item["descriptions"].append(accessory["detail_final"])
            if accessory["origin"] == "manual":
                item["manual"] += quantity
            elif accessory["status"] == "modified":
                item["winmaker_modified"] += quantity
            else:
                item["winmaker_original"] += quantity

    result: list[dict[str, Any]] = []
    for code, item in grouped.items():
        has_manual = item["manual"] != 0
        has_winmaker = item["winmaker_original"] != 0 or item["winmaker_modified"] != 0
        if has_manual and has_winmaker:
            origin = "Mixto"
        elif has_manual:
            origin = "Agregado Manualmente"
        elif item["winmaker_modified"] != 0:
            origin = "Winmaker Modificado"
        else:
            origin = "Winmaker"
        descriptions = item["descriptions"]
        final_description = resolutions.get(code, {}).get("final_description")
        result.append(
            {
                "Código": code,
                "Cantidad Total": int(item["Cantidad Total"])
                if item["Cantidad Total"] == item["Cantidad Total"].to_integral()
                else float(item["Cantidad Total"]),
                "Detalle": final_description or (descriptions[0] if len(descriptions) == 1 else ""),
                "Origen": origin,
                "Conflicto": len(descriptions) > 1 and not final_description,
                "Descripciones detectadas": descriptions,
                "Subtotales": {
                    "Winmaker": float(item["winmaker_original"]),
                    "Winmaker Modificado": float(item["winmaker_modified"]),
                    "Agregado Manualmente": float(item["manual"]),
                },
                "Observaciones": " | ".join(
                    dict.fromkeys(
                        accessory.get("observations", "")
                        for typology in document["typologies"]
                        for accessory in typology["accessories"]
                        if not accessory["excluded"]
                        and accessory["code"] == code
                        and accessory.get("observations")
                    )
                ),
            }
        )
    return sorted(result, key=lambda item: _natural_key(item["Código"]))


def resolve_description(
    session: Session,
    order: OrderRecord,
    code: str,
    final_description: str,
) -> None:
    if not final_description.strip():
        raise DomainError("La descripción final es obligatoria.")
    document = load_document(order)
    document.setdefault("conflict_resolutions", {})[code] = {
        "final_description": final_description.strip(),
        "resolved_at": datetime.now(timezone.utc).isoformat(),
    }
    add_audit(
        session,
        order,
        "description_conflict_resolved",
        entity_type="consolidated_accessory",
        entity_id=code,
        field="description",
        final=final_description.strip(),
    )
    save_document(order, document)


def finalize_order(session: Session, order: OrderRecord) -> None:
    document = load_document(order)
    pending = [item for item in document["typologies"] if item["review"]["status"] != "confirmed"]
    if pending:
        raise DomainError(f"Quedan {len(pending)} tipologías sin confirmar.")
    conflicts = [item for item in consolidated_accessories(order) if item["Conflicto"]]
    if conflicts:
        raise DomainError(f"Quedan {len(conflicts)} conflictos de descripción sin revisar.")
    order.status = "confirmed"
    add_audit(session, order, "order_finalized")


def audit_events(session: Session, order_id: str) -> list[dict[str, Any]]:
    records = (
        session.query(AuditEventRecord)
        .filter(AuditEventRecord.order_id == order_id)
        .order_by(AuditEventRecord.id.desc())
        .all()
    )
    return [
        {
            "id": record.id,
            "event_type": record.event_type,
            "page": record.page,
            "field": record.field,
            "original_value": record.original_value,
            "previous_value": record.previous_value,
            "final_value": record.final_value,
            "created_at": record.created_at.isoformat(),
        }
        for record in records
    ]


def make_xlsx(order: OrderRecord, kind: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    if kind == "glasses":
        sheet.title = "Vidrios"
        rows = glass_rows(order)
    elif kind == "accessories":
        sheet.title = "Accesorios"
        rows = [
            {
                key: value
                for key, value in row.items()
                if key in ("Código", "Cantidad Total", "Detalle", "Origen", "Observaciones")
            }
            for row in consolidated_accessories(order)
        ]
    else:
        raise DomainError("Tipo de exportación no válido.")
    _fill_sheet(sheet, rows)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _fill_sheet(sheet: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        sheet.append(["Sin datos"])
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24352F")
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 55)
        sheet.column_dimensions[column[0].column_letter].width = width
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def make_csv(order: OrderRecord, kind: str) -> bytes:
    import csv

    rows = glass_rows(order) if kind == "glasses" else [
        {
            key: value
            for key, value in row.items()
            if key in ("Código", "Cantidad Total", "Detalle", "Origen", "Observaciones")
        }
        for row in consolidated_accessories(order)
    ]
    output = io.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()), delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def make_pdf(order: OrderRecord, kind: str) -> bytes:
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    if kind == "glasses":
        title = "Listado de Vidrios"
        rows = glass_rows(order)
    elif kind == "accessories":
        title = "Listado de Accesorios"
        rows = [
            {
                key: value
                for key, value in row.items()
                if key in ("Código", "Cantidad Total", "Detalle", "Origen", "Observaciones")
            }
            for row in consolidated_accessories(order)
        ]
    else:
        raise DomainError("Tipo de exportación no válido.")
    title_style = ParagraphStyle(
        "BrandTitle",
        parent=styles["Title"],
        textColor=colors.HexColor("#123A58"),
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        spaceAfter=6 * mm,
    )
    story: list[Any] = [
        Paragraph(f"{title} - {order.client_name}", title_style),
    ]
    story.append(_pdf_table(rows, styles))
    document.build(story)
    return output.getvalue()


def _pdf_table(rows: list[dict[str, Any]], styles: Any) -> Table:
    if not rows:
        return Table([["Sin datos"]])
    headers = list(rows[0].keys())
    header_style = ParagraphStyle(
        "TableHeader",
        parent=styles["BodyText"],
        textColor=colors.white,
        fontName="Helvetica-Bold",
        fontSize=7,
        leading=8,
    )
    body_style = ParagraphStyle(
        "TableBody",
        parent=styles["BodyText"],
        fontSize=7,
        leading=8.5,
        textColor=colors.HexColor("#15344A"),
    )
    data = [[Paragraph(str(header), header_style) for header in headers]]
    for row in rows:
        data.append([Paragraph(str(row.get(header, "")), body_style) for header in headers])
    if headers == ["Código", "Cantidad Total", "Detalle", "Origen", "Observaciones"]:
        column_widths = [20 * mm, 27 * mm, 69 * mm, 31 * mm, 39 * mm]
    elif len(headers) == 8:
        column_widths = [21 * mm, 16 * mm, 24 * mm, 17 * mm, 18 * mm, 24 * mm, 40 * mm, 26 * mm]
    else:
        column_widths = None
    table = Table(data, repeatRows=1, hAlign="LEFT", colWidths=column_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123A58")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B7C9D2")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF5F7")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table
